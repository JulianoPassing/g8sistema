# -*- coding: utf-8 -*-
"""Lê tabelas Pantaneiro 5% e 7% (xlsx) e atualiza JSON + HTML embutidos."""
import json
import os
import re
import sys

try:
    import openpyxl
except ImportError:
    print("Instale openpyxl: pip install openpyxl", file=sys.stderr)
    raise

DEFAULT_XLSX_5 = (
    r"g:\Outros computadores\Meu laptop\G8 Representações\PANTANEIRO"
    + r"\Tabela de Preços Completa (281400) -  Região Sulsudeste - Comissão 5%.xlsx"
)
DEFAULT_XLSX_7 = (
    r"g:\Outros computadores\Meu laptop\G8 Representações\PANTANEIRO"
    + r"\Tabela de Preços Completa (301400) -  Região Sulsudeste - Comissão 7%.xlsx"
)

SKIP_CATEGORY_PREFIXES = (
    "REF.",
    "POLITICA",
    "PEDIDO",
    "PRIMEIRA",
    "ACRESCIMO",
    "DESCONTO",
    "CORTA VENTO E BOTAS",
    "BOTAS DE COURO FATURAMENTO",
    "OBEDECER",
    "ACIMA DE 3",
    "COMPRAS",
    "FATURAMENTO VIA",
)

CAT_MAP = {
    "AVENTURA": "Aventura",
    "ESTILO": "Estilo",
    "URBANO": "Urbano",
    "PRO": "PRO",
    "CORTA VENTO": "Corta Vento",
    "URBANO - ACESSORIOS": "Acessórios",
    "URBANO - ACESSÓRIOS": "Acessórios",
    "ACESSORIOS": "Acessórios",
    "ACESSÓRIOS": "Acessórios",
}

SIZE_TOKENS = {"PP", "P", "M", "G", "GG", "EX", "EXG", "2G", "3G", "4G", "5G"}

HTML_TARGETS = {
    5: [
        "pantaneiro5.html",
        "b2b-pantaneiro5.html",
    ],
    7: [
        "pantaneiro7.html",
        "b2b-pantaneiro7.html",
    ],
}

MARKER_START = "// ========== INÍCIO DA LISTA DE PRODUTOS ATUALIZADA =========="
MARKER_END = "// ========== FIM DA LISTA DE PRODUTOS ATUALIZADA =========="


def parse_price(value):
    if value is None:
        return None
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return round(float(value), 2)
    try:
        return round(float(str(value).replace(",", ".")), 2)
    except ValueError:
        return None


def parse_ref(value):
    if value is None:
        return None
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        if float(value) == int(value):
            return str(int(value))
        return str(value).strip()
    text = str(value).strip()
    return text or None


def normalize_category(raw):
    key = " ".join(str(raw).split()).upper()
    return CAT_MAP.get(key, str(raw).strip())


def is_category_row(ref, code, desc, price):
    if not ref or code or desc or price is not None:
        return False
    up = ref.upper()
    if up in ("REF.",):
        return False
    if any(up.startswith(p) for p in SKIP_CATEGORY_PREFIXES):
        return False
    if len(up) > 30:
        return False
    return True


def parse_tamanhos(raw):
    if raw is None or str(raw).strip() == "":
        return ["UN"]
    text = str(raw).strip()
    upper = text.upper()

    # Botas por par de numeração: 35/36-37/38-39/40...
    size_pairs = re.findall(r"\d+/\d+", text)
    if len(size_pairs) >= 2:
        return size_pairs
    if len(size_pairs) == 1 and "-" in text:
        return size_pairs

    if "BOTAS" in upper and (
        "/" in text or "UNICO" in upper or "ÚNICO" in upper or "P - EX" in text
    ):
        return [text.replace("UNICO", "ÚNICO").replace("Unico", "ÚNICO")]

    parts = [p.strip() for p in re.split(r"\s*-\s*", text) if p.strip()]
    if len(parts) > 1 and all(
        p in SIZE_TOKENS or re.match(r"^\d{1,2}$", p) for p in parts
    ):
        return parts

    return [text]


def parse_xlsx(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    items = []
    category = None

    for row in range(1, ws.max_row + 1):
        ref_raw, code, desc_raw, tam_raw, price_raw = (
            ws.cell(row, col).value for col in range(1, 6)
        )
        ref = parse_ref(ref_raw)
        desc = " ".join(str(desc_raw).split()) if desc_raw else ""
        price = parse_price(price_raw)

        if is_category_row(ref, code, desc, price):
            category = normalize_category(ref)
            continue

        if not desc or not ref:
            continue

        if price is None:
            continue

        items.append(
            {
                "CATEGORIA": category or "Sem categoria",
                "REFERENCIA": ref,
                "DESCRIÇÃO": desc,
                "TAMANHOS": parse_tamanhos(tam_raw),
                "PRECO": price,
            }
        )

    return items


def js_string(value):
    return json.dumps(value, ensure_ascii=False)


def format_tamanhos_js(tamanhos):
    if len(tamanhos) <= 1:
        return f"[{js_string(tamanhos[0])}]"
    inner = ",\n      ".join(js_string(t) for t in tamanhos)
    return f"[\n      {inner},\n    ]"


def to_js_array(items):
    lines = ["window.produtosData = ["]
    for index, item in enumerate(items):
        comma = "," if index < len(items) - 1 else ""
        tamanhos_js = format_tamanhos_js(item["TAMANHOS"])
        preco = item["PRECO"]
        preco_js = str(int(preco)) if preco == int(preco) else f"{preco:.2f}"
        block = (
            f"  {{\n"
            f"    CATEGORIA: {js_string(item['CATEGORIA'])},\n"
            f"    REFERENCIA: {js_string(item['REFERENCIA'])},\n"
            f"    DESCRIÇÃO: {js_string(item['DESCRIÇÃO'])},\n"
            f"    TAMANHOS: {tamanhos_js},\n"
            f"    PRECO: {preco_js},\n"
            f"  }}{comma}"
        )
        lines.append(block)
    lines.append("];")
    return "\n".join(lines)


def patch_html(html_path, items):
    with open(html_path, encoding="utf-8") as f:
        html = f.read()
    start = html.find(MARKER_START)
    end = html.find(MARKER_END)
    if start == -1 or end == -1:
        raise ValueError(f"Marcadores não encontrados em {html_path}")
    end += len(MARKER_END)
    replacement = (
        f"{MARKER_START}\n"
        f"{to_js_array(items)}\n"
        f"{MARKER_END}"
    )
    with open(html_path, "w", encoding="utf-8", newline="\n") as f:
        f.write(html[:start] + replacement + html[end:])


def write_json(path, items):
    with open(path, "w", encoding="utf-8", newline="\n") as f:
        json.dump(items, f, ensure_ascii=False, indent=2)
        f.write("\n")


def main():
    root = os.path.join(os.path.dirname(__file__), "..")
    public_dir = os.path.join(root, "public")
    xlsx5 = sys.argv[1] if len(sys.argv) > 1 else DEFAULT_XLSX_5
    xlsx7 = sys.argv[2] if len(sys.argv) > 2 else DEFAULT_XLSX_7

    results = []
    for comissao, path in ((5, xlsx5), (7, xlsx7)):
        if not os.path.isfile(path):
            print(f"Arquivo não encontrado: {path}", file=sys.stderr)
            sys.exit(1)
        items = parse_xlsx(path)
        json_path = os.path.join(public_dir, f"produtos-pantaneiro{comissao}.json")
        write_json(json_path, items)
        for filename in HTML_TARGETS[comissao]:
            patch_html(os.path.join(public_dir, filename), items)
        results.append(
            {
                "comissao": comissao,
                "total": len(items),
                "json": json_path,
                "html": HTML_TARGETS[comissao],
                "fonte": path,
                "categorias": sorted({item["CATEGORIA"] for item in items}),
            }
        )
        print(f"Pantaneiro {comissao}%: {len(items)} produtos -> {json_path}")

    print(json.dumps(results, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
