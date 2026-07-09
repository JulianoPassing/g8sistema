# -*- coding: utf-8 -*-
"""Lê tabelas BKB Atacado/Distribuidor/Especial (xlsx) e gera public/bkb-fabrica-produtos.js."""
from __future__ import annotations

import json
import re
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("Instale openpyxl: pip install openpyxl", file=sys.stderr)
    raise

ROOT = Path(__file__).resolve().parents[1]
OUT_JS = ROOT / "public" / "bkb-fabrica-produtos.js"

DEFAULT_BASE = Path(r"g:\Outros computadores\Meu laptop\G8 Representações\BKB")
DEFAULT_FILES = {
    "atacado": DEFAULT_BASE / "Tabela Atacado BKB - Janeiro 2026 G8.xlsx",
    "distribuidor": DEFAULT_BASE / "Tabela Distribuidor BKB - Janeiro 2026 G8.xlsx",
    "especial": DEFAULT_BASE / "Tabela Especial BKB - Janeiro 2026 G8.xlsx",
}

REF_RE = re.compile(r"^[A-Z]{2,5}-[\w]+$", re.I)

# Bloco de categorias preservado do arquivo atual (enriquecimento em runtime).
CATEGORIA_FOOTER = r'''
(function bkbFabricaEnriquecerCategorias() {
  var rules = [
    ["GRNT", "Kit guarnição tampa de válvula"],
    ["GCB", "Guarnição carburador"],
    ["GPB", "Guarda pó bengala"],
    ["CXM", "Coxim de coroa"],
    ["RBG", "Kit retentor de bengala + guarda pó"],
    ["RBE", "Retentor de bengala"],
    ["PTR", "Pedaleira"],
    ["PTD", "Pedaleira"],
    ["PDI", "Pedaleira"],
    ["PDL", "Borracha de estribo"],
    ["RVV", "Retentor haste de válvula"],
    ["RTT", "Retentor tampa de válvula"],
    ["RPF", "Vedação pinça de freio"],
    ["PRB", "Presilha de rabeta"],
    ["MGC", "Mangueira de combustível"],
    ["GTL", "Guarnição tampa lateral"],
    ["GRN", "Guarnição tampa de válvula"],
    ["VMP", "Vedação motor de partida"],
    ["VIE", "Vedação injeção eletrônica"],
    ["BUJ", "Bujão / parafuso cárter"],
    ["BUC", "Bujão / parafuso cárter"],
    ["BQE", "Bucha quadro elástico"],
    ["BAM", "Kit borracha amortecedor"],
    ["ARR", "Arruelas de alumínio"],
    ["ANE", "Anel do escape"]
  ];
  rules.sort(function (a, b) {
    return b[0].length - a[0].length;
  });

  function categoriaPorRef(ref) {
    var r = String(ref || "").toUpperCase();
    for (var i = 0; i < rules.length; i++) {
      if (r.indexOf(rules[i][0]) === 0) return rules[i][1];
    }
    return "Demais peças";
  }

  window.BKB_FABRICA_CATEGORIA_ORDEM = [
    "Anel do escape",
    "Arruelas de alumínio",
    "Kit borracha amortecedor",
    "Bucha quadro elástico",
    "Bujão / parafuso cárter",
    "Coxim de coroa",
    "Guarnição carburador",
    "Guarda pó bengala",
    "Guarnição tampa de válvula",
    "Kit guarnição tampa de válvula",
    "Guarnição tampa lateral",
    "Mangueira de combustível",
    "Borracha de estribo",
    "Pedaleira",
    "Presilha de rabeta",
    "Retentor de bengala",
    "Kit retentor de bengala + guarda pó",
    "Vedação pinça de freio",
    "Retentor tampa de válvula",
    "Retentor haste de válvula",
    "Vedação injeção eletrônica",
    "Vedação motor de partida",
    "Demais peças"
  ];

  if (!window.produtosData || !Array.isArray(window.produtosData)) return;
  window.produtosData = window.produtosData.map(function (p) {
    var c = p.CATEGORIA || categoriaPorRef(p.REF);
    var copy = {};
    for (var k in p) {
      if (Object.prototype.hasOwnProperty.call(p, k)) copy[k] = p[k];
    }
    copy.CATEGORIA = c;
    return copy;
  });
})();
'''


def parse_price(value):
    if value is None:
        return None
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return round(float(value), 2)
    m = re.search(r"(\d+[.,]\d+|\d+)", str(value).replace("R$", ""))
    if not m:
        return None
    return round(float(m.group(1).replace(",", ".")), 2)


def parse_sheet(path: Path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    items = {}
    order = []
    for r in range(1, ws.max_row + 1):
        ref_raw = ws.cell(r, 1).value
        if ref_raw is None:
            continue
        ref = str(ref_raw).strip()
        if not REF_RE.match(ref):
            continue
        desc = ws.cell(r, 2).value
        price = parse_price(ws.cell(r, 7).value)
        items[ref] = {
            "REF": ref,
            "MODELO": str(desc).strip() if desc else None,
            "PRECO": price,
        }
        order.append(ref)
    return items, order


def merge_tables(files: dict[str, Path]):
    tables = {}
    orders = {}
    for tipo, path in files.items():
        if not path.exists():
            raise FileNotFoundError(path)
        items, order = parse_sheet(path)
        tables[tipo] = items
        orders[tipo] = order
        print(f"{tipo}: {len(items)} itens ({path.name})")

    all_refs = []
    seen = set()
    for tipo in ("atacado", "distribuidor", "especial"):
        for ref in orders[tipo]:
            if ref not in seen:
                seen.add(ref)
                all_refs.append(ref)

    merged = []
    warnings = []
    for ref in all_refs:
        a = tables["atacado"].get(ref, {})
        d = tables["distribuidor"].get(ref, {})
        e = tables["especial"].get(ref, {})
        modelo = a.get("MODELO") or e.get("MODELO") or d.get("MODELO") or ""
        # Descrição mais completa quando especial traz texto maior (ex.: RVV)
        for cand in (e.get("MODELO"), a.get("MODELO"), d.get("MODELO")):
            if cand and len(cand) > len(modelo):
                modelo = cand

        precos = {
            "atacado": a.get("PRECO"),
            "distribuidor": d.get("PRECO"),
            "especial": e.get("PRECO"),
        }
        for k, src in (
            ("atacado", a),
            ("distribuidor", d),
            ("especial", e),
        ):
            if not src:
                warnings.append(f"{ref}: ausente na tabela {k}")
            elif precos[k] is None:
                warnings.append(f"{ref}: preço {k} sem valor no Excel")

        merged.append({"REF": ref, "MODELO": modelo, "PRECOS": precos})

    return merged, warnings


def format_number(n):
    if n is None:
        return "null"
    if isinstance(n, (int, float)) and not isinstance(n, bool):
        n = float(n)
        if abs(n - round(n)) < 1e-9:
            return str(int(round(n)))
        s = f"{n:.10f}".rstrip("0").rstrip(".")
        return s
    return "null"


def to_js(items):
    lines = ["window.produtosData = ["]
    for i, p in enumerate(items):
        comma = "," if i < len(items) - 1 else ""
        precos = p["PRECOS"]
        lines.append("  {")
        lines.append(f'    "REF": {json.dumps(p["REF"], ensure_ascii=False)},')
        lines.append(f'    "MODELO": {json.dumps(p["MODELO"], ensure_ascii=False)},')
        lines.append('    "PRECOS": {')
        lines.append(f'      "atacado": {format_number(precos.get("atacado"))},')
        lines.append(f'      "distribuidor": {format_number(precos.get("distribuidor"))},')
        lines.append(f'      "especial": {format_number(precos.get("especial"))}')
        lines.append("    }")
        lines.append(f"  }}{comma}")
    lines.append("];")
    lines.append("")
    lines.append(CATEGORIA_FOOTER.lstrip("\n"))
    return "\n".join(lines)


def main():
    files = dict(DEFAULT_FILES)
    if len(sys.argv) >= 4:
        files = {
            "atacado": Path(sys.argv[1]),
            "distribuidor": Path(sys.argv[2]),
            "especial": Path(sys.argv[3]),
        }

    items, warnings = merge_tables(files)
    OUT_JS.write_text(to_js(items), encoding="utf-8")
    print(f"\nGerado {OUT_JS} com {len(items)} produtos")
    if warnings:
        print("Avisos:")
        for w in warnings:
            print(f"  - {w}")


if __name__ == "__main__":
    main()
